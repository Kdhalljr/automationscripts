# Script to parse computer list (shown below) for use in python automation scripts


from openpyxl import load_workbook


# Spreadsheet Location
COMPUTER_INVENTORY_SPREADSHEET = 'xxxxxx'


# Different groupings of columns pertain to different machine types
LOCAL_MACHINES_COL = 9
MGR_MACHINES_COL = 10


# Helper Functions
def find_row_range(ws):
    return range(2, ws.max_row+1)


def find_col_range(ws):
    return range(1, ws.max_column+1)



def find_store_pos_cols(ws):
    ASCII_SUBTRACT = 64
    first, last = ws.merged_cell_ranges[0].split(':')
    return range(ord(first[:-1]) - ASCII_SUBTRACT, ord(last[:-1]) - ASCII_SUBTRACT + 1)


def find_store_buy_cols(ws):
    ASCII_SUBTRACT = 64
    first, last = ws.merged_cell_ranges[1].split(':')
    return range(ord(first[:-1]) - ASCII_SUBTRACT, ord(last[:-1]) - ASCII_SUBTRACT +1)


def remove_duplicates(mylist):
    newlist = []
                 
    for item in mylist:
        if item not in newlist:
            newlist.append(item)

    return newlist


def search_ws_by_col(ws):
    comp_list = []
                 
    for col in find_col_range(ws):
        for row in find_row_range(ws):
            if (ws.cell(column=col, row=row).value != 'n/a'
            and ws.cell(column=col, row=row).value != None):
                comp_list.append(ws.cell(column=col, row=row).value.encode('ascii'))

    return comp_list


def search_col_range(ws, col_range):
    comp_list = []
                 
    if (str(type(col_range)) != "<type 'list'>"):
        col_range = [col_range]
                 
    for row in find_row_range(ws):
        for col in col_range:
            if (ws.cell(column=col, row=row).value != 'n/a'
            and ws.cell(column=col, row=row).value != None):
                comp_list.append(ws.cell(column=col, row=row).value.encode('ascii'))

    return comp_list


# Build List Functions
def build_cro_pos_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, find_store_pos_cols(ws))


def build_f5_pos_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, find_store_pos_cols(ws))


def build_cro_buy_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, find_store_buy_cols(ws))


def build_f5_buy_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, find_store_buy_cols(ws))


def build_cro_local_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LOCAL_MACHINES_COL)


def build_f5_local_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    comp_list = search_col_range(ws, LOCAL_MACHINES_COL)
    return search_col_range(ws, LOCAL_MACHINES_COL)


def build_cro_mgr_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, MGR_MACHINES_COL)


def build_f5_mgr_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, MGR_MACHINES_COL)


def build_hq01_comp_list():
    LIST_COL = 1
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LIST_COL)


def build_hq01_laptop_list():
    LIST_COL = 2
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LIST_COL)


def build_hq01_test_machine_list():
    LIST_COL = 3
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LIST_COL)


def build_dc99_comp_list():
    LIST_COL = 1
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LIST_COL)


def build_dc99_buy_list():
    LIST_COL = 2
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LIST_COL)


def build_dc99_ship_comp_list():
    LIST_COL = 3
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LIST_COL)


def build_db78_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_ws_by_col(ws)


def build_db79_list():
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_ws_by_col(ws)


def build_mgmt_comp_list():
    LIST_COL = 1
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LIST_COL)


def build_mgmt_surface_list():
    LIST_COL = 2
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LIST_COL)


def build_mgmt_laptop_list():
    LIST_COL = 3
    wb = load_workbook(COMPUTER_INVENTORY_SPREADSHEET)
    ws = wb['xxxxxx']
    return search_col_range(ws, LIST_COL)


def build_pos_list():
    return remove_duplicates(build_cro_pos_list() + build_f5_pos_list())


def build_buy_list():
    return remove_duplicates(build_cro_buy_list() + build_f5_buy_list())


def build_local_list():
    return remove_duplicates(build_cro_local_list() + build_f5_local_list())


def build_mgr_list():
    return remove_duplicates(build_cro_mgr_list() + build_f5_mgr_list())


def build_cro_list():
    return remove_duplicates(build_cro_pos_list() + build_cro_buy_list() + build_cro_local_list() + build_cro_mgr_list())


def build_f5_list():
    return remove_duplicates(build_f5_pos_list() + build_f5_buy_list() + build_f5_local_list() + build_f5_mgr_list())


def build_stores_list():
    return remove_duplicates(build_cro_list() + build_f5_list())


def build_hq01_list():
    return remove_duplicates(build_hq01_comp_list() + build_hq01_laptop_list() + build_hq01_test_machine_list())


def build_dc99_list():
    return remove_duplicates(build_dc99_comp_list() + build_dc99_buy_list() + build_dc99_ship_comp_list())


def build_db_list():
    return remove_duplicates(build_db78_list() + build_db79_list())


def build_mgmt_list():
    return remove_duplicates(build_mgmt_comp_list() + build_mgmt_surface_list() + build_mgmt_laptop_list())


def build_computer_list():
    return remove_duplicates(build_stores_list() + build_hq01_list() + build_dc99_list() + build_db_list() + build_mgmt_list())
